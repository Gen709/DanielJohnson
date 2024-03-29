import csv
import openpyxl
from openpyxl.utils import get_column_letter

from django.contrib.auth.models import User
from django.utils import timezone
from student.models import Student
from school.models import Group
from datetime import datetime

class ExtractStudent():
    def __init__(self, request):
        self.request = request
        self.file_name = request.FILES['csv_file'] 

    def get_csv_reader(self):
        #TODO: mettre les entete en all cap
        reader = []
        # with open(self.file_name, mode='r', newline='', encoding='utf-8-sig') as file:
        r = csv.DictReader(self.file_name.read().decode('utf-8-sig').splitlines())
        for row in r:
            capitalized_row = {key.upper(): value for key, value in row.items()}
            reader.append(capitalized_row)
                
        return reader         
            
    def get_excel_reader(self):
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook.active
        reader = []
        header = [cell.value for cell in sheet[1]]
        # print(header)
        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            for idx, cell in enumerate(row):
                if header[idx]:
                    row_data[header[idx].upper()] = cell.value
                else:
                    row_data[header[idx]] = " "
            reader.append(row_data)
            
        return reader

    def get_reader(self):
        
        if self.file_name.name.endswith('.csv'):
            reader = self.get_csv_reader()
        elif self.file_name.name.endswith('.xlsx'):
            reader = self.get_excel_reader()
        return reader

    def get_student_data_dict_list(self, reader):
        
        all_student_data_dict_list = []

        for row in reader:
            
            if row['FICHE']:
                # print(row['FICHE'])
                nom_prenom = row.get('NOM', '').split(',')
                nom = nom_prenom[0].strip() if nom_prenom else ''
                prenom = nom_prenom[1].strip() if len(nom_prenom) > 1 else ''
                
                plan_intervention_value = row.get("PLAN D'INTERVENTION", None)
                if plan_intervention_value:
                    plan_intervention_value = plan_intervention_value.strip().lower()

                # Handle GROUPE REPERE field
                group_repere_name = row.get('GROUPE', None)
                if group_repere_name:
                    group_repere_name = group_repere_name.strip()
                    groupe_repere, create = Group.objects.get_or_create(nom=group_repere_name)
                    # groupe_repere = groupe_repere.strip()

                if nom == "Ben Younes":
                    print(f"--------- In the get_student_data_dict_list BEFORE the info is stored in the dict for {prenom} the groupe_repere is {groupe_repere}")
                    print(f"---------  {prenom} the groupe_repere is {row.get('FICHE', '')}")

                # Map CSV fields to model fields
                student_data = {"nom": nom, 
                                "prenom": prenom, 
                                "comite_clinique": False, 
                                "date_ref_comite_clinique": None, 
                                "plan_intervention": plan_intervention_value == 'oui', 
                                "groupe_repere": groupe_repere, 
                                "code": None, 
                                "fiche": row.get('FICHE', ''), 
                                "etat_situation": None, 
                                "dob": row.get('DATE DE NAISSANCE', None), 
                                "lang": row.get('LANGUE PARLÉE À LA MAISON', ''), 
                                "is_student": True,  # Set to True for new students, 
                                "created_by": self.request.user, 
                                "date_created": timezone.now(), 
                                "date_is_student_changed": None
                                }
                
                # if student_data.get("nom") == "Ben Younes":
                #     print(f"--------- In the get_student_data_dict_list AFTER the info has been stored in the dict for {prenom} the groupe_repere is {student_data.get('groupe_repere')}")
                
                # print(student_data)
                all_student_data_dict_list.append(student_data)

        return all_student_data_dict_list

    def set_student_to_inactive(self, gv_all_new_student_list):
        all_students_qs = Student.objects.all()
        # gv = General Vanier
        gv_new_student_fiche_list = [student_data["fiche"] for student_data in gv_all_new_student_list]
        # students in the db that are no longuer in the extract
        old_student = all_students_qs.exclude(fiche__in=gv_new_student_fiche_list)
        # print("total number: ", len(all_students_qs), "Old numbers: ",len(old_student))
        # set those student is_student status to False
        number_of_inactivated_records = old_student.update(is_student=False, date_is_student_changed=timezone.now())
        return number_of_inactivated_records

    def update_and_create_students(self, gv_all_new_student_list):
        existing_active_student_in_db_fiche_list = [s.fiche for s in Student.objects.filter(is_student=1)]
        existing_updated_student_list = []
        newly_created_student_list = []
        for student_data in gv_all_new_student_list:
            if student_data['fiche'] in existing_active_student_in_db_fiche_list:
                existing_student = Student.objects.get(fiche=student_data['fiche'])
                for field, value in student_data.items():
                    if getattr(existing_student, field) != value and field in ['classification', 'plan_intervention', 'groupe_repere', 'lang']:
                        setattr(existing_student, field, value)
                        setattr(existing_student, "date_is_student_cha  nged", timezone.now())
                        existing_student.save()
                        existing_updated_student_list.append(existing_student)
            else:
                student = Student.objects.update_or_create(
                                                            fiche=student_data.pop('fiche', None),
                                                            defaults=student_data  # Use defaults instead of **
                                                        )
                newly_created_student_list.append(student)

        return newly_created_student_list, existing_updated_student_list

    def update_data(self):
        reader = self.get_reader()
        gv_all_new_student_list = self.get_student_data_dict_list(reader)
        number_of_inactivated_records = self.set_student_to_inactive(gv_all_new_student_list)
        newly_created_student_list, existing_updated_student_list = self.update_and_create_students(gv_all_new_student_list)
        return {"number_of_inactivated_records": number_of_inactivated_records, 
                "newly_created_student_list": newly_created_student_list,
                "existing_updated_student_list": existing_updated_student_list,
                "number_newly_created_student":len(newly_created_student_list),
                "number_existing_updated_student":len(existing_updated_student_list)}
    
    def update_student_group(self):
        reader = self.get_reader()
        gv_all_new_student_list = self.get_student_data_dict_list(reader)
        # inactive_student = Student.objects.filter(is_student=False)
        # inactive_student.delete()
        for student in gv_all_new_student_list:
            fiche = student.get('fiche')
            nom = student.get('nom')
            prenom = student.get("prenom")
            groupe = student.get('groupe_repere')
            try:
                student_obj = Student.objects.get(nom=nom, prenom=prenom)
                student_obj.groupe_repere = groupe
                student_obj.fiche = fiche
                student_obj.save()

                # if student_obj.nom == "Ben Younes":
                #     print(f"--------- In the update_student_group AFTER saving the information for {student_obj.prenom} the groupe_repere is {student_obj.groupe_repere}")
                #     print("After Update - Student Object:", student_obj)
                #     print("After Update - Groupe Repere:", student_obj.groupe_repere)

                # Add more print statements to debug
                # print("After update - Groupe Repere:", student_obj.groupe_repere)
            except Student.DoesNotExist:
                # print(f"Student with fiche {fiche} does not exist.")
                pass

        return {"status": "Update complete"}


def get_student_grade_dict(evaluation_qs):
    data_dict = {}
    for e in evaluation_qs:
        if e.etape in data_dict.keys():
            if e.competence_evaluee.matiere.description in data_dict[e.etape].keys():
                if e.competence_evaluee.competence.nom in data_dict[e.etape][e.competence_evaluee.matiere.description].keys():
                    pass
                else:
                    data_dict[e.etape][e.competence_evaluee.matiere.description][e.competence_evaluee.competence.nom] = e.note
            else:
                data_dict[e.etape][e.competence_evaluee.matiere.description ] = {}
        else:
            data_dict[e.etape] = {}      
    
    return data_dict

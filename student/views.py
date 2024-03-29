from django.shortcuts import render, redirect
from django.db.models import Count, Q
from urllib.parse import unquote
from django.http.response import JsonResponse, HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required

from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.core.paginator import Paginator
from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags

from .models import Student, StatusAction, Problematique, StatusProblematique, Action,\
    ActionSuggestion, CodeEtudiant, Grades, EtatDeLaSituation, Evaluation
from problematiques.models import Item
from school.models import Group, CompetencesEvaluees, Matiere, Competence
from teacher.models import Professional, RegularTeacher
from .forms import CSVUploadForm, EtatDeLaSituationForm
from .util import ExtractStudent, get_student_grade_dict
from datetime import datetime as dt, timedelta

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle, Color, Fill
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter

import re

# Create your views here.

class EvaluationViewList(ListView):
    model=Evaluation
    context_object_name = 'evaluation_list'
    template_name = 'student/evaluation_list.html'
    paginate_by = 25

    def get_queryset(self):
        return Evaluation.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        paginator = Paginator(self.get_queryset(), self.paginate_by)
        page = self.request.GET.get('page')
        context['evaluation_list'] = paginator.get_page(page)
        return context


def test_eta_de_la_situation(request, id):
    if request.user.is_authenticated:
        intervenant = User.objects.get(id=request.user.id)
        student = Student.objects.get(id=id)
        creation_date = dt.strptime(request.POST.get("creationDateStr", dt.strftime(dt.today().date(),  "%b. %d, %Y")), "%b. %d, %Y").date()
        id = request.POST.get("id")
        try:
            # Check if an EtatDeLaSituation instance exist with the given creator and date exists
            etat_instance = EtatDeLaSituation.objects.get(id = id)

            # etat_instance = EtatDeLaSituation.objects.get(creation_date = creation_date, creator=intervenant)
            # Determine if it's a new instance or an existing one
            is_new_instance = False
        except:
            is_new_instance = True
        try:
            # Check if an EtatDeLaSituation instance exist with the given creator and date exists
            etat_instance = EtatDeLaSituation.objects.get(creator = intervenant, creation_date=dt.today().date())
            # Determine if it's a new instance or an existing one
            is_new_instance = False
        except:
            is_new_instance = True
        


        if request.method == 'POST':
            # etat_de_la_situation_id = request.POST.get('etat_de_la_situation_id')
            etat_situation_str = request.POST.get('etatdelasituation')
            # Now you can use 'is_new_instance' to differentiate between new and existing instances
            if is_new_instance:
                # pass
                # It's a new instance
                # Perform actions for a new instance
                EtatDeLaSituation.objects.create(student = student, 
                                                creator = intervenant, 
                                                text = etat_situation_str)
            else:
                # It's an existing instance
                # Perform actions for an existing instance
                # Update the attributes of the instance
                etat_instance.text=etat_situation_str
                etat_instance.modifier=intervenant
                etat_instance.save() # Modify this with the new text
                

            print("******************************************************************")
            print("New instance ?",is_new_instance)
            print(" / by:", intervenant.get_full_name(), 'intervenant:', intervenant)
            print(" / date:", creation_date, "/ id:", id)
            print("str", etat_situation_str)
            print("******************************************************************")


            latest_object = EtatDeLaSituation.objects.latest('id')
            next_id = latest_object.id + 1
            older_record_qs = EtatDeLaSituation.objects.filter(student=student)
            context = {'intervenant': intervenant,
                    'student': student,
                    'older_record': older_record_qs,
                    'next_id': next_id,
                    'is_new_instance': is_new_instance}
            
            return render(request, "student/test_detail_etat_de_la_situation.html", context)
        else:
            try:
                latest_object = EtatDeLaSituation.objects.latest('id')
                next_id = latest_object.id + 1
            except:
                next_id = 1

            older_record_qs = EtatDeLaSituation.objects.filter(student=student)
            context = {'intervenant': intervenant,
                    'student': student,
                    'older_record': older_record_qs,
                    'next_id': next_id,
                    'is_new_instance': is_new_instance}
            return render(request, "student/test_detail_etat_de_la_situation.html", context)
    

@csrf_exempt
def ajax_student_problematique_action_detail_update(request):
    if request.method == "POST":
        a = Action.objects.get(id=int(request.POST.get('action_id', "").split("_")[2]))
        detail = request.POST.get('action_desc', "")

        a.detail = detail
        a.save()

        return redirect(a.problematique.eleve.get_absolute_url())


@csrf_exempt
def ajax_student_problematique_action_status_update(request):
    if request.method == "POST":
        a = Action.objects.get(id=int(request.POST.get('action_id', "").split("_")[2]))
        s = StatusAction.objects.get(id=request.POST.get('status_id', ""))

        a.status = s
        a.save()

        return redirect(a.problematique.eleve.get_absolute_url())


@csrf_exempt
def ajax_student_problematique_update(request):
    if request.method == "POST":
        problematique_id = request.POST.get("problematique_id", "No problematique id")
        p = Problematique.objects.get(id=problematique_id)
        problematique_desc = request.POST.get("problematique_desc", "No problematique desc")
        p.detail = problematique_desc
        p.save()
        return redirect(p.eleve.get_absolute_url())


def ajax_search_probleme_action_sugestions(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # input_str = request.GET.get('term', '')
        input_str = unquote(request.GET['term'])
        data_list = [{x.id: {'description': x.nom}} for x in
                     ActionSuggestion.objects.filter(nom__istartswith=input_str)]

        data = JsonResponse(data_list, safe=False)

        mimetype = 'application/json'

        if len(data_list) > 0 and input_str != '':
            return HttpResponse(data, mimetype)
        else:
            data = JsonResponse([{'description': 'Aucune suggestion'}], safe=False)
            return HttpResponse(data, mimetype)

@csrf_exempt
def ajax_note_student(request):
    student_id = request.POST.get("student_id")
    data_dict = {}
    # etudiant
    s = Student.objects.get(id=student_id)
    fiche = s.fiche
    # niveau de l'étudiant
    niveau = s.groupe_repere.classification # s.group.classification
    data_dict["niveau"] = niveau.nom
    # notes de l'année en cour
    # n = Grades.objects.filter(student_id=student_id)
    n = Grades.objects.filter(no_fiche=fiche)
    # liste des notes de l'élève
    grades_list = [x.note for x in n]
    data_dict["grade_list"] = grades_list
    # liste des cours de l'élève
    class_name_list = [x.matiere for x in n]
    classification_className_tuple_list = [(x.classification, x.matiere) for x in n]
    data_dict["class_name_list"] = class_name_list
    data_dict["classification_className_tuple_list"] = classification_className_tuple_list
    # Notes de groupe
    data_dict["class_grades"] = {}
    for classification_className_tuple in data_dict["classification_className_tuple_list"]:
        data_dict["class_grades"][classification_className_tuple[1]] = [x.note for x in Grades.objects.filter(
            matiere=classification_className_tuple[1])]


    data = JsonResponse(data_dict, safe=False)
    mimetype = 'application/json'
    return HttpResponse(data, mimetype)

@csrf_exempt
def ajax_note_student_2(request):
    student_id = request.POST.get("student_id")
    semester = int(request.POST.get("semester"))
    data_dict = {}
    data_dict["semester"] = semester
    s = Student.objects.get(id=student_id)
    classe = s.groupe_repere
    evaluation_qs = Evaluation.objects.filter(etudiant=s, competence_evaluee__matiere__matiere_de_base=True, etape=semester).order_by("competence_evaluee__matiere__subject_code", "competence_evaluee__competence__nom")
    data_dict["grade_list"] = [evaluation.note for evaluation in evaluation_qs if evaluation.etape == semester]
    data_dict["class_name_list"] = [evaluation.competence_evaluee.matiere.description + " " +  evaluation.competence_evaluee.competence.nom 
                                    for evaluation in evaluation_qs if evaluation.etape == semester]

    data_dict["class_grades"] = {evaluation.competence_evaluee.matiere.description + " " + evaluation.competence_evaluee.competence.nom: [x.note for x in Evaluation.objects.filter(competence_evaluee = evaluation.competence_evaluee, 
                                                                                                                                                                                    etudiant__groupe_repere = classe,
                                                                                                                                                                                    etape=semester)] for evaluation in evaluation_qs}

    data = JsonResponse(data_dict, safe=False)
    mimetype = 'application/json'
    return HttpResponse(data, mimetype)


def ajax_search_student(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # input_str = request.GET.get('term', '')
        input_str = unquote(request.GET['term'])
        data_list = [{x.id: {'description': x.nom + " " + x.prenom + " - Group: " + x.groupe_repere.nom,
                             'url': x.get_absolute_url()}} for x in Student.objects.filter(nom__istartswith=input_str, is_student=True)]

        data = JsonResponse(data_list, safe=False)

        mimetype = 'application/json'
        if len(data_list) > 0 and input_str != '':
            return HttpResponse(data, mimetype)
        else:
            data = JsonResponse([{'description': "Nothing with the term " + input_str,
                                  'url': "NO URL"}], safe=False)
            return HttpResponse(data, mimetype)


def ajax_update_student(request):
    parameter = unquote(request.GET.get('param'))
    value = unquote(request.GET.get('value'))
    student_id = unquote(request.GET.get('student_id'))

    s = Student.objects.get(pk=student_id)

    if parameter == "comite_cliniqueflexSwitchCheck":
        s.comite_clinique = value
    elif parameter == "plan_interventionRadioOptionsflexSwitchCheck":
        s.plan_intervention = value
    elif parameter == "student_code":
        c = CodeEtudiant.objects.get(pk=value)
        s.code = c

    s.save()

    return redirect('student-detail', pk=student_id)


@login_required
def student_detail_view_2(request, pk):
    if request.user.is_authenticated:
        intervenant = User.objects.get(id=request.user.id)
        # staff = User.objects.get(username=request.user.get_username())
        student = Student.objects.get(id=pk)
        evaluation_qs = Evaluation.objects.filter(etudiant=student, competence_evaluee__matiere__matiere_de_base=True).order_by("competence_evaluee__matiere__subject_code", "competence_evaluee__competence__nom")
        grades = get_student_grade_dict(evaluation_qs)
        semester_set = {evaluation.etape for evaluation in evaluation_qs}
        
        latest_semester = max(semester_set) if semester_set else 1
        # faire arriver à la derniere evaluation. allerchercher l'étape la plus elevé en excluant 4 et 8
        pros = Professional.objects.all().order_by('speciality')
        teacher = RegularTeacher.objects.filter(group__id=student.groupe_repere.id)
        direction = User.objects.filter(id=student.groupe_repere.classification.owner.id)
        responsable_qs_dict = {'professionals':pros, 
                                "regularteacher":teacher, 
                                "direction":direction
                                }
        
        problematiques = Item.objects.all()
        statusproblematique = StatusProblematique.objects.all()
        statusaction = StatusAction.objects.all()
        code_etudiant = CodeEtudiant.objects.all()

        older_record_qs = EtatDeLaSituation.objects.filter(student=student)

        etat_instance = None

        try:
            EtatDeLaSituation.objects.latest('id')
            latest_object = EtatDeLaSituation.objects.latest('id')
            next_id = latest_object.id + 1
        except:
            next_id = 1

        try:
            # has the intervenant already posted totady
            etat_instance = EtatDeLaSituation.objects.get(student=student, creator=intervenant, creation_date=dt.today().date())
            print('etat_instance', etat_instance)
            if etat_instance:
                is_new_instance = False
            else:
                is_new_instance = True
        except:
            is_new_instance = True

        # print("******************** Is new instance", is_new_instance, "request method", request.method)
        # print("******************** Student", student, "number of records", len(older_record_qs))
        # print("******************** today's records", older_record_qs.filter(student=student, creator=intervenant, creation_date=dt.today().date()).text)

        context = {'student': student, 
                    'grades': grades, 
                    'latest_semester': latest_semester,
                    'semester_set':semester_set,
                    'evaluation_qs':evaluation_qs,
                    'intervenant': intervenant,
                    'problematiques': problematiques, 
                    'statusaction': statusaction, 
                    'statusproblematique': statusproblematique, 
                    'responsable_qs': responsable_qs_dict, 
                    'code_etudiant': code_etudiant, 
                    'older_record_qs': older_record_qs,
                    'is_new_instance': is_new_instance,
                    'next_id':next_id
                    }
        
        if request.method == 'POST':
            post_id = request.POST.get("post_id")
            etat_situation_str = request.POST.get('etatdelasituation')

            print("******************** Is new instance", is_new_instance, "request method", request.method)
            print("******************** Student", student, "number of records", len(older_record_qs))
            print("******************** today's records for this student by this intervenant", older_record_qs.filter(student=student, creator=intervenant, creation_date=dt.today().date()))
            
            try:
                etat_instance = EtatDeLaSituation.objects.get(id=post_id)
                is_new_instance = False
            except:
                pass

            if not is_new_instance:
                # It's an existing instance
                # Perform actions for an existing instance
                # Update the attributes of the instance
                etat_instance.text=etat_situation_str
                etat_instance.modifier=intervenant
                etat_instance.save() # Modify this with the new text
            else:
                # maybe it's not a new instance but an old instance
                
                print(etat_instance)
                print(etat_situation_str)
                if etat_instance:
                    etat_instance.text=etat_situation_str
                    etat_instance.modifier=intervenant
                    etat_instance.save()
                else:
                    EtatDeLaSituation.objects.create(student = student, 
                                                    creator = intervenant, 
                                                    text = etat_situation_str)
                

            # print("******************************************************************")
            # print("New instance ?",is_new_instance)
            # print(" / by:", intervenant.get_full_name(), 'intervenant:', intervenant)
            # # print(" / date:", creation_date, "/ id:", id)
            # print("str", etat_situation_str)
            # print("******************************************************************")

            context["older_record_qs"] = EtatDeLaSituation.objects.filter(student=student)
            latest_object = EtatDeLaSituation.objects.latest('id')
            context["next_id"] = latest_object.id + 1

            # context = {'student': student, 
            #             # 'staff': staff, 
            #             'intervenant': intervenant,
            #             'problematiques': problematiques, 
            #             'statusaction': statusaction, 
            #             'statusproblematique': statusproblematique, 
            #             'responsable_qs': responsable_qs_dict, 
            #             'code_etudiant': code_etudiant, 
            #             # 'etat_de_la_situation': etat_de_la_situation, 
            #             'older_record_qs': older_record_qs,
            #             'is_new_instance': is_new_instance,
            #             'next_id': next_id
            #         }

            return render(request, 'student/detail_3.html', context)
        else:
            return render(request, 'student/detail_3.html', context)


@login_required
def student_detail_view(request, pk):
    # a revoir
    staff = User.objects.get(username=request.user.get_username())
    student = Student.objects.get(pk=pk)
    # etat_de_la_situation = EtatDeLaSituation.objects.filter(student=student)
    # responsable_qs = User.objects.all()
    pros = Professional.objects.all().order_by('speciality')
    teacher = RegularTeacher.objects.filter(group__id=student.groupe_repere.id)
    direction = User.objects.filter(id=student.groupe_repere.classification.owner.id)
    responsable_qs_dict = {'professionals':pros, 
                           "regularteacher":teacher, 
                           "direction":direction}
    
    problematiques = Item.objects.all()
    statusproblematique = StatusProblematique.objects.all()
    statusaction = StatusAction.objects.all()
    code_etudiant = CodeEtudiant.objects.all()

    context = {'student': student,
               'staff': staff,
               'problematiques': problematiques,
               'statusaction': statusaction,
               'statusproblematique': statusproblematique,
               'responsable_qs': responsable_qs_dict,
               'code_etudiant': code_etudiant,
            #    'etat_de_la_situation': etat_de_la_situation
               }

    return render(request, 'student/detail.html', context)


@login_required
def student_problematique_create_view(request):
    student = Student.objects.get(pk=request.POST.get("eleve_id"))
    instigateur = User.objects.get(pk=request.POST.get("staff_id"))
    problematique = Item.objects.get(pk=request.POST.get("problematique_id"))
    detail = request.POST.get("prob_detail")
    status = StatusProblematique.objects.get(pk=request.POST.get("status_prob"))

    p = Problematique(nom=problematique,
                      status=status,
                      instigateur=instigateur,
                      detail=detail,
                      eleve=student)
    p.save()

    return redirect(student.get_absolute_url())


@csrf_exempt
def student_problematique_update_status(request):
    p = Problematique.objects.get(id=int(request.POST.get('problematique_id', "").split("_")[2])) # probleme
    s = StatusProblematique.objects.get(id=request.POST.get('status_id', ""))
    p.status = s
    p.save()

    return redirect(p.eleve.get_absolute_url())


@login_required
def student_action_problematique_insert_view(request):
    # TODO: staff is created 
    student = Student.objects.get(pk=request.POST.get("eleve_id"))
    staff = User.objects.get(pk=request.POST.get("staff_id"))
    problematique = Problematique.objects.get(pk=request.POST.get("problematique_id"))
    description = request.POST.get("description")
    detail = request.POST.get("detail")
    action_status = StatusAction.objects.get(pk=request.POST.get("action_status_id"))
    responsable = User.objects.get(pk=request.POST.get("responsable_id"))

    a = Action(createur=staff,
               responsable=responsable,
               description=description,
               detail=detail,
               status=action_status,
               problematique=problematique)
    a.save()

    return redirect(student.get_absolute_url())


@login_required
def comitecliniquestudentlistview(request):
    # student_comite_clinique_dict = {
    #     classification: [s for s in Student.objects.filter(groupe_repere__classification=classification).filter(comite_clinique=True)]
    #     for classification in
    #     {c for c in Classification.objects.filter(group__student__comite_clinique=True)}
    # }

    student_comite_clinique_dict = {
        groupe: [s for s in Student.objects.filter(groupe_repere=groupe).filter(comite_clinique=True).order_by('nom')]
        for groupe in
        {g for g in Group.objects.filter(student__comite_clinique=True).order_by('nom')}
    }

    context = {'student_comite_clinique_dict': student_comite_clinique_dict}

    return render(request, "student/comite_clinique_accordeon.html", context)
    # return render(request, "student/comite_clinique_list.html", context)


@login_required
def cours_ete_list_view(request):
    context = {}
    return render(request, 'student/coursdete_eleve.html', context)


@login_required
def eleve_evaluation_list(request):
    suggestion_list = ActionSuggestion.objects.all()
    a = Action.objects.filter(description__in=[x.nom for x in suggestion_list], problematique__eleve__is_student=True).exclude(status__id=4)

    context = {'evaluation_suggestion_list': suggestion_list,
               'evaluation_liste': a
               }

    return render(request, 'student/evaluation_eleve.html', context)


@login_required
def eleve_problematique_list_view(request):

    code_etudiant_qs = CodeEtudiant.objects.exclude(code=0).annotate(num_student=Count('student', filter=Q(student__is_student=True)))

    results = {code: {classification:
                          [student for student in code.student_set.filter(is_student=True, groupe_repere__classification=classification.id)]
                      for classification in {etudiant.groupe_repere.classification for etudiant in code.student_set.filter(is_student=True)}
                      }
               for code in code_etudiant_qs
               }
    # classification_list = problematique_list.student.classification.distinct()

    # classification = CodeEtudiant.objects.
    context = {"results": results
               # "classification_list":test
               }

    return render(request, 'student/problemes_eleve.html', context)


@login_required
def download_excel_data(request, user_id=None):

    def cleanhtml(raw_html):
        CLEANR = re.compile('<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});')

        if raw_html != "" and (type(raw_html) == str):
            cleantext = re.sub(CLEANR, '', raw_html)
            return cleantext

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Comité_Clinique_Daniel_Johnson_' + dt.now().strftime(
        "%Y-%m-%d %H:%M:%S") + ".xlsx"
    # cell.alignment = Alignment(vertical='center')
    wb = Workbook()
    # small text
    v_centerAlignment = Alignment(horizontal='center', 
                                  vertical='center', 
                                  text_rotation=0, 
                                  wrap_text=False, 
                                  shrink_to_fit=False, 
                                  indent=0)
    # large text
    v_centerAlignment_large = Alignment(horizontal='left', 
                                        vertical='top', 
                                        text_rotation=0, 
                                        wrap_text=True, 
                                        shrink_to_fit=False, 
                                        indent=0)

    title_centerAlignment = Alignment(horizontal='center', 
                                      vertical='center', 
                                      text_rotation=0, 
                                      wrap_text=False, 
                                      shrink_to_fit=False, 
                                      indent=0)
    # c_white = Color('00FFFFFF')


    bigTitleStyle = NamedStyle(name="bigTitleStyle") 
    bigTitleFont = Font(name='Calibri', 
                     size=22, 
                     bold=True, 
                     italic=False, 
                     vertAlign=None, 
                     underline='none', 
                     strike=False, 
                     color='00000000')
    bigTitleStyle.font = bigTitleFont
    bd = Side(style='thick', color="00C0C0C0") 
    bigTitleStyle.border = Border(left=None, top=bd, right=bd, bottom=bd)
    bigTitleFill = PatternFill(fill_type="solid", 
                                start_color='0099CCFF', 
                                end_color='0099CCFF')
    bigTitleStyle.fill = bigTitleFill
    bigTitleStyle.alignment = title_centerAlignment

    secondTitleStyle = NamedStyle(name="secondTitleStyle") 
    secondTitleFont = Font(name='Calibri', 
                           size=16, 
                           bold=False, 
                           italic=False, 
                           vertAlign=None, 
                           underline='none', 
                           strike=False, 
                           color='00000000')
    secondTitleStyle.font = secondTitleFont
    secondTitleStyle.alignment = title_centerAlignment
    secondTitleFill = PatternFill(fill_type="solid", 
                                  start_color='0099CCFF', 
                                  end_color='0099CCFF')
    secondTitleStyle.fill = secondTitleFill
    SecondTitlebd = Side(style="thin", color="00C0C0C0")
    secondTitleStyle.border = Border(left=SecondTitlebd, top=SecondTitlebd, right=SecondTitlebd, bottom=SecondTitlebd)

    ws1 = wb.active

    column_width_list = [('H', 45), ('M', 45)]

    for column_letter, width in column_width_list:
        ws1.column_dimensions[column_letter].width = width


    ws1.title = "Comité Clinique"
    ws1['A1'] = "Étudiant"
    ws1['A1'].style = bigTitleStyle
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws1['J1'] = "Problématiques"
    ws1['J1'].style = bigTitleStyle
    ws1.merge_cells(start_row=1, start_column=10, end_row=1, end_column=13)
    ws1['N1'] = "Action"
    ws1['N1'].style = bigTitleStyle
    ws1.merge_cells(start_row=1, start_column=14, end_row=1, end_column=18)

    c = ws1['A3']
    ws1.freeze_panes = c
    ws1['A2'] = "Fiche"
    ws1['A2'].style = secondTitleStyle
    ws1['B2'] = "Nom"
    ws1['B2'].style = secondTitleStyle
    ws1['C2'] = "Prenom"
    ws1['C2'].style = secondTitleStyle
    ws1['D2'] = "Groupe Repere"
    ws1['D2'].style = secondTitleStyle
    ws1['E2'] = "Classification"
    ws1['E2'].style = secondTitleStyle
    ws1['F2'] = "Comite Clinique"
    ws1['F2'].style = secondTitleStyle
    ws1['g2'] = "Plan Intervention"
    ws1['g2'].style = secondTitleStyle
    ws1['h2'] = "État Situation"
    ws1['h2'].style = secondTitleStyle
    ws1['i2'] = "Classification"
    ws1['i2'].style = secondTitleStyle
    ws1['j2'] = "Nom"
    ws1['j2'].style = secondTitleStyle
    ws1['k2'] = "Status"
    ws1['k2'].style = secondTitleStyle
    ws1['l2'] = "Instigateur"
    ws1['l2'].style = secondTitleStyle
    ws1['m2'] = "Detail"
    ws1['m2'].style = secondTitleStyle
    ws1['n2'] = "Createur"
    ws1['n2'].style = secondTitleStyle
    ws1['o2'] = "Responsable"
    ws1['o2'].style = secondTitleStyle
    ws1['p2'] = "Description"
    ws1['p2'].style = secondTitleStyle
    ws1['q2'] = "Detail"
    ws1['q2'].style = secondTitleStyle
    ws1['r2'] = "Status"
    ws1['r2'].style = secondTitleStyle

    i = 3

   

    for student in Student.objects.filter(comite_clinique=True).order_by('groupe_repere', 'nom'):
        if student.problematique_set.all():
            for problematique in student.problematique_set.all():
                if problematique.action_set.all():
                    for action in problematique.action_set.all():
                        ws1['A' + str(i)] = student.fiche
                        ws1['A' + str(i)].alignment = v_centerAlignment
                        ws1['B' + str(i)] = student.nom
                        ws1['B' + str(i)].alignment = v_centerAlignment
                        ws1['C' + str(i)] = student.prenom
                        ws1['C' + str(i)].alignment = v_centerAlignment
                        ws1['D' + str(i)] = student.groupe_repere.nom if student.groupe_repere else None
                        ws1['D' + str(i)].alignment = v_centerAlignment
                        ws1['E' + str(i)] = student.groupe_repere.classification.nom if student.groupe_repere else None
                        ws1['E' + str(i)].alignment = v_centerAlignment
                        ws1['F' + str(i)] = student.comite_clinique
                        ws1['F' + str(i)].alignment = v_centerAlignment
                        ws1['g' + str(i)] = student.plan_intervention
                        ws1['g' + str(i)].alignment = v_centerAlignment
                        etat_de_la_situation_str = " ".join([x.text.replace("</p>", "\n") for x in student.etatdelasituation_set.all()])
                        ws1['h' + str(i)] = cleanhtml(etat_de_la_situation_str) 
                        ws1['h' + str(i)].alignment = v_centerAlignment_large
                        ws1['i' + str(i)] = student.groupe_repere.classification.nom if student.groupe_repere else None
                        ws1['i' + str(i)].alignment = v_centerAlignment

                        ws1['j' + str(i)] = problematique.nom.nom
                        ws1['j' + str(i)].alignment = v_centerAlignment
                        ws1['k' + str(i)] = problematique.status.nom
                        ws1['k' + str(i)].alignment = v_centerAlignment
                        ws1['l' + str(i)] = problematique.instigateur.first_name if problematique.instigateur else None
                        ws1['l' + str(i)].alignment = v_centerAlignment
                        ws1['m' + str(i)] = cleanhtml(problematique.detail)
                        ws1['m' + str(i)].alignment = v_centerAlignment_large

                        ws1['n' + str(i)] = action.createur.first_name if action.createur else None
                        ws1['n' + str(i)].alignment = v_centerAlignment
                        ws1['o' + str(i)] = action.responsable.first_name if action.responsable else None
                        ws1['o' + str(i)].alignment = v_centerAlignment
                        ws1['p' + str(i)] = action.description
                        ws1['p' + str(i)].alignment = v_centerAlignment_large
                        ws1['q' + str(i)] = cleanhtml(action.detail)
                        ws1['q' + str(i)].alignment = v_centerAlignment_large
                        ws1['r' + str(i)] = action.status.nom
                        ws1['r' + str(i)].alignment = v_centerAlignment
                        i += 1
                else:
                    ws1['A' + str(i)] = student.fiche
                    ws1['A' + str(i)].alignment = v_centerAlignment
                    ws1['B' + str(i)] = student.nom
                    ws1['B' + str(i)].alignment = v_centerAlignment
                    ws1['C' + str(i)] = student.prenom
                    ws1['C' + str(i)].alignment = v_centerAlignment
                    ws1['D' + str(i)] = student.groupe_repere.nom if student.groupe_repere else None
                    ws1['D' + str(i)].alignment = v_centerAlignment
                    ws1['E' + str(i)] = student.groupe_repere.classification.nom if student.groupe_repere else None
                    ws1['E' + str(i)].alignment = v_centerAlignment
                    ws1['F' + str(i)] = student.comite_clinique
                    ws1['F' + str(i)].alignment = v_centerAlignment
                    ws1['g' + str(i)] = student.plan_intervention
                    ws1['g' + str(i)].alignment = v_centerAlignment
                    etat_de_la_situation_str = " ".join([x.text.replace("</p>", "\n") for x in student.etatdelasituation_set.all()])
                    ws1['h' + str(i)] = cleanhtml(etat_de_la_situation_str) 
                    # ws1['h' + str(i)] = cleanhtml(student.etat_situation)
                    ws1['h' + str(i)].alignment = v_centerAlignment_large
                    ws1['i' + str(i)] = student.groupe_repere.classification.nom if student.groupe_repere else None
                    ws1['i' + str(i)].alignment = v_centerAlignment

                    ws1['j' + str(i)] = problematique.nom.nom
                    ws1['j' + str(i)].alignment = v_centerAlignment
                    ws1['k' + str(i)] = problematique.status.nom
                    ws1['k' + str(i)].alignment = v_centerAlignment
                    ws1['l' + str(i)] = problematique.instigateur.first_name if problematique.instigateur else None
                    ws1['l' + str(i)].alignment = v_centerAlignment
                    ws1['m' + str(i)] = cleanhtml(problematique.detail)
                    ws1['m' + str(i)].alignment = v_centerAlignment_large
                    i += 1

        else:
            ws1['A' + str(i)] = student.fiche
            ws1['A' + str(i)].alignment = v_centerAlignment
            ws1['B' + str(i)] = student.nom
            ws1['B' + str(i)].alignment = v_centerAlignment
            ws1['C' + str(i)] = student.prenom
            ws1['C' + str(i)].alignment = v_centerAlignment
            ws1['D' + str(i)] = student.groupe_repere.nom if student.groupe_repere else None
            ws1['D' + str(i)].alignment = v_centerAlignment
            ws1['E' + str(i)] = student.groupe_repere.classification.nom if student.groupe_repere else None
            ws1['E' + str(i)].alignment = v_centerAlignment
            ws1['F' + str(i)] = student.comite_clinique
            ws1['F' + str(i)].alignment = v_centerAlignment
            ws1['g' + str(i)] = student.plan_intervention
            ws1['g' + str(i)].alignment = v_centerAlignment
            etat_de_la_situation_str = " ".join([x.text.replace("</p>", "\n") for x in student.etatdelasituation_set.all()])
            ws1['h' + str(i)] = cleanhtml(etat_de_la_situation_str) 
            # ws1['h' + str(i)] = cleanhtml(student.etat_situation)
            ws1['h' + str(i)].alignment = v_centerAlignment_large
            ws1['i' + str(i)] = student.groupe_repere.classification.nom if student.groupe_repere else None
            ws1['i' + str(i)].alignment = v_centerAlignment
            i += 1

    
    def fit_col_width(ws1):
        column_widths = []
        for row in ws1:
            for i, cell in enumerate(row):
                if cell.value is not None and cell.row == 2:
                    if len(column_widths) > i:
                        if len(cell.value) > column_widths[i]:
                            column_widths[i] = len(cell.value)
                    else:
                        column_widths += [len(cell.value)]

        for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
            ws1.column_dimensions[get_column_letter(i)].width = column_width * 2

    fit_col_width(ws1)

    ws1.auto_filter.ref = "A2:R"+str(i)
    import io

    # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename=Comité_Clinique_Daniel_Johnson_' + dt.now().strftime(
    #     "%Y-%m-%d %H:%M:%S")+'.xlsx'
    # wb.save(filename=response)

    # Save the workbook to an in-memory binary stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create the HttpResponse object with the appropriate content type and headers
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Comité_Clinique_Daniel_Johnson_' + dt.now().strftime(
        "%Y-%m-%d %H:%M:%S") + '.xlsx'

    # Return the response
    return response
    

@login_required
def upload_csv(request):
    if request.method == 'POST':
        action = request.POST.get('update_groupe', None)
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            es = ExtractStudent(request)
            if action:
                context = es.update_student_group()
            else:
                context = es.update_data()

            return render(request, 'summary_page.html', context)
    else:
        form = CSVUploadForm()
    return render(request, 'upload_csv.html', {'form': form})


@login_required
def upload_summary(request):
    new_students = Student.objects.filter(date_created__gte=timezone.now() - timedelta(seconds=5))
    inactive_students = Student.objects.filter(is_student=False)

    return render(request, 'summary_page.html', {'new_students': new_students, 'inactive_students': inactive_students})


@login_required
def review_changes(request):
    if request.method == 'POST':
        # Process the selected changes (approve)

        # Example: Print the approved changes for demonstration
        print("Approved changes:")
        print("New students to create:", request.POST.getlist('new_students'))
        print("Students to update:", request.POST.getlist('inactive_students'))

        # You can process the approved changes as needed

        return redirect('upload_csv')  # Redirect to the CSV upload page after approval

    else:
        return HttpResponse("GET request to the review_changes page is not supported.")


def send_email(request):
    # retrieve the etat de la situation that changed whitin a certain time frame:

    # Retrieve mailing list (you would replace this with your own logic to fetch the mailing list)
    recipients = ['recipient1@example.com', 'recipient2@example.com']

    # Retrieve content (you would replace this with your own logic to fetch content)
    email_content = 'This is the email content.'

    # Generate the email content
    email_subject = 'Your Subject Here'
    email_from = 'your-email@example.com'

    # Create the email message using a template (optional)
    html_content = render_to_string('student/comite_clinique_email_update.html', {'content': email_content})
    text_content = strip_tags(html_content)

    # Create the email message
    email = EmailMultiAlternatives(email_subject, text_content, email_from, recipients)
    # email = EmailMessage(email_subject, text_content, email_from, recipients)
    email.attach_alternative(html_content, "text/html")

    # Send the email
    email.send()

    return render(request, 'student/comite_clinique_email_update.html')
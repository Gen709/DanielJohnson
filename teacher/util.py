import csv
import string

import openpyxl
from openpyxl.utils import get_column_letter

from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.utils import timezone

from .models import RegularTeacher, Professional, SpecialtyTeacher
from school.models import Group, Classification
from datetime import datetime

def is_csv_empty(file):
    try:
        # Read the CSV file and check if it's empty
        csv_reader = csv.reader(file.read().decode('utf-8').splitlines())
        for row in csv_reader:
            if any(field for field in row if field.strip()):
                return False  # CSV is not empty
        return True  # CSV is empty
    except csv.Error:
        return True  # Invalid CSV format or empty file

class ExtractTeacher():
    def __init__(self, request):
        self.request = request
        self.file_name = request.FILES['csv_file']
        self.hashed_password = make_password("General-Vanier")

    def get_csv_reader(self):
        # reader = []
        # # with open(self.file_name, mode='r', newline='', encoding='utf-8-sig') as file:
        # print(f"????????????????????????????File Content: {self.file_name.read()}")
        # r = csv.DictReader(self.file_name.read().decode('utf-8').splitlines())
        # for row in r:
        #     cleaned_row = {key.replace('\ufeff', ''): value for key, value in row.items()}
        #     capitalized_row = {key.upper(): value for key, value in cleaned_row.items()}
        #     print(capitalized_row)
        #     reader.append(capitalized_row) 
        # print("****************************** CSV READER ALERT", reader)      
        # return reader 
      
        try:
            # Reset file pointer to the beginning of the file
            self.file_name.seek(0)

            # Read the CSV content and decode assuming UTF-8 encoding
            csv_content = self.file_name.read().decode('utf-8')

            # Print the content to debug
            print(f"File Content: {csv_content}")
            l = []
            # Use csv.DictReader with the content
            reader = csv.DictReader(csv_content.splitlines())
            for row in reader:
                cleaned_row = {key.replace('\ufeff', ''): value for key, value in row.items()}
                capitalized_row = {key.upper(): value for key, value in cleaned_row.items()}
                print(capitalized_row)
                l.append(capitalized_row)
            # Reset file pointer again to the beginning of the file for any further use
            self.file_name.seek(0)

            return l
        
        except Exception as e:
            print(f"Error: {str(e)}")
            return []  # Return an empty list in case of an error
              
   
    def get_excel_reader(self):
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook.active
        reader = []
        header = [cell.value for cell in sheet[1]]
        print(header)
        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            for idx, cell in enumerate(row):
                if header[idx]:
                    row_data[header[idx].upper()] = cell.value
                else:
                    row_data[header[idx]] = " "
            reader.append(row_data)
            
        return reader

    def get_reader(self):
        if self.file_name.name.endswith('.csv'):
            print("************csv reader selected****************************")
            reader = self.get_csv_reader()
        elif self.file_name.name.endswith('.xlsx'):
            reader = self.get_excel_reader()
        return reader
    
    def identify_extract(self):
        # TODO: need to add the email when available
        reader = self.get_reader()
        print(reader)
        regular_teacher_extract_row_list = ['GROUPE', 'NOM', 'LOCAUX', 'MATIERES']
        professional_extract_row_list = ['PRENOM NOMFAMILLE', 'LOCAL', 'SPECIALITÉ']
        specialty_extract_row_list = ["TYPE", "GROUPE", "NOM", "LOCAUX", "SPÉCIALITÉ"]
        
        this_extract_row_list = [k for k in reader[0].keys()]
        
        if  this_extract_row_list == regular_teacher_extract_row_list:
            return "regular_teacher_extract"
        elif this_extract_row_list == professional_extract_row_list:
            return "professional_extract"
        elif this_extract_row_list == specialty_extract_row_list:
            return "specialist_extract"

    def create_or_update_group(self):
        reader = self.get_reader()
        created_or_updated_group_list = []
        classification = Classification.objects.all()
        # print(classification)
        group_name_list = [teacher_dict["GROUPE"] for teacher_dict in reader if '-' not in teacher_dict["GROUPE"]]
        for classi in classification:
            # print(classi)
            for group_name in group_name_list:
                # print(group_name)
                # print(classi.nom, len(classi.nom))
                if len(classi.nom) == 1:
                    # print(string.ascii_lowercase.index(classi.nom.lower()), int(group_name[1])) # classification de GV
                    if string.ascii_lowercase.index(classi.nom.lower()) == int(group_name[1]):
                        # print(classi.nom, group_name, int(group_name[1]))
                        g, created = Group.objects.update_or_create(nom=group_name, 
                                                                    defaults={"classification":classi})
                        created_or_updated_group_list.append((g, created))

        return created_or_updated_group_list

    def create_or_update_regular_teacher(self):
        #TODO: Utiliser le courriel pour la mise a jour ou la création
        reader = self.get_reader()
        operation_list = []
        # plain_text_password = "General-Vanier"
        # hashed_password = make_password(plain_text_password)
        for row in reader:
            prenom_nomfamille = row["NOM"].split()
            prenom = prenom_nomfamille[0]
            nom_famille = " ".join(prenom_nomfamille[1:])
            try:
                g = Group.objects.get(nom=row["GROUPE"])
            except:
                g = None
            try:
                courriel = row["COURRIEL"]
            except:
                courriel = ""
            # print(prenom_nomfamille, prenom, nom_famille, g)   
            teacher_obj, created = RegularTeacher.objects.update_or_create(username = prenom[:3]+nom_famille[:3],
                                                                            defaults = {"password": self.hashed_password, 
                                                                                        "first_name": prenom, 
                                                                                        "last_name": nom_famille, 
                                                                                        "email": courriel, 
                                                                                        "group": g, 
                                                                                        "matière": row["MATIERES"]})
            operation_list.append((teacher_obj, created))

        return operation_list
    
    def create_or_update_professionals(self):
        reader = self.get_reader()
        professional_list = []
        for row in reader:
            prenom_nomfamille = row['PRENOM NOMFAMILLE'].split()
            prenom = prenom_nomfamille[0]
            nom_famille = " ".join(prenom_nomfamille[1:])
            local = row['LOCAL' ]
            speciality = row['SPECIALITÉ']
            try:
                courriel = row["COURRIEL"]
            except:
                courriel = ""

            pro_dict = {"password": self.hashed_password,
                        "first_name": prenom,  
                        "last_name": nom_famille,  
                        "email": "",  
                        "local": local,  
                        "speciality": speciality}
            a, b = Professional.objects.update_or_create(username = prenom[:3]+nom_famille[:3],
                                                         defaults = pro_dict)
            professional_list.append((a, b))

        return professional_list
      
    def create_or_update_specialist(self):
        reader = self.get_reader()
        # plain_text_password = "General-Vanier"
        # hashed_password = make_password(plain_text_password)
        specialist_list = []
        for row in reader:
            prenom_nomfamille = row['NOM'].split()
            prenom = prenom_nomfamille[0]
            nom_famille = " ".join(prenom_nomfamille[1:])
            local = row['LOCAUX' ]
            speciality = row['SPÉCIALITÉ']

            try:
                courriel = row["COURRIEL"]
            except:
                courriel = ""

            speciality_dict = {"password": self.hashed_password, 
                               "first_name": prenom, 
                               "last_name": nom_famille, 
                               "email": courriel, 
                               "matière": speciality,
                               "local": local
                              }

            specialist, created = SpecialtyTeacher.objects.update_or_create(username = prenom[:3]+nom_famille[:3],
                                                                            defaults = speciality_dict)
            specialist_list.append((specialist, created))
            
        return specialist_list
        
    def update_regularteacher_data(self):
        reader = self.get_reader()
        updated_or_created_group_obj_list = self.create_or_update_group()
        updated_or_created_regular_teacher_obj_list = self.create_or_update_regular_teacher()
        context = {"updated_or_created_group_obj_list": updated_or_created_group_obj_list, 
                   "updated_or_created_regular_teacher_obj_list": updated_or_created_regular_teacher_obj_list, 
                  }
        
        return context
    
    def update_professional_data(self):
        return self.create_or_update_professionals()
    
    def update_specialist_data(self):
        return self.create_or_update_specialist()
    
    def update_data(self):
        reader = self.get_reader()
        if self.identify_extract() == "regular_teacher_extract":
            return self.update_regularteacher_data()
        elif self.identify_extract() == "professional_extract":
            return self.update_professional_data()
        elif self.identify_extract() == "specialist_extract":
            return self.update_specialist_data()

import openpyxl
from .models import Matiere, Competence, CompetencesEvaluees
from student.models import Student, Evaluation
import os
import io


matiere_headings_list = ["École", "Matière", "Description", "Catégorie", "Visible aux parents/élèves", "Mat 2ième", None, None ]
objectif_matiere_headings_list = ["Fiche", "Nom de l'élève", "Classe", "Matière", "Grp", "Obj. - Matière", "No", "Description", "Obj. - Résultat final", "Pond", "[1]", "[2]", "[3]", "[4]", "[5]", "[6]", "[7]", "[8]", "Remarque", None, None]
local_heading_list = ["École",	"Local",	"Description	Champ 01",	'Champ 02',	'Capacité',	'Contenu',	"Remarque",	"Date maj. hor-cal", "Date maj. hor-cyc", "Conflit accepté", "Loc hors Eco"]
cours_heading_list = ["École", "Cours",	"Description", "Grille", "Type", "Grp.prév."]

def get_headings(file_path):
     # Load the Excel workbook
    heading_list = []
    workbook = openpyxl.load_workbook(io.BytesIO(file_path))
    # workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    first_row_values = sheet[1]
    for cell in first_row_values:
        heading_list.append(cell.value)
    return heading_list


def handle_uploaded_file(file):
    file_path = os.path.join('uploads', file.name)
    with open(file_path, 'wb') as destination:
        for chunk in file.chunks():
            destination.write(chunk)
    return file_path


def upload_data_from_excel_matiere(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(io.BytesIO(file_path))
    # workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active    
    # Iterate through rows and create new instances
    for row in sheet.iter_rows(min_row=2, values_only=True):
        mat, created = Matiere.objects.update_or_create(school_code=row[0], 
                                                            subject_code=row[1], 
                                                            description=row[2], 
                                                            category=row[3], 
                                                            visible_to_parents=row[4], 
                                                            mat_2nd=row[5]
                                                            )
        print(f"Matiere : {mat} has been created : {created}")
    
    matieres = Matiere.objects.all()
    for matiere in matieres:
        if matiere.subject_code[:3] == "MAT" or matiere.subject_code[:3] == "ANG" or matiere.subject_code[:3] == "FRA":
            matiere.matiere_de_base = True
            matiere.save()


def upload_data_from_excel_competences(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(io.BytesIO(file_path))
    # workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Iterate through rows and create new instances
    for row in sheet.iter_rows(min_row=2, values_only=True):
        comp, c_created = Competence.objects.update_or_create(nom = row[7].strip())
        matiere = Matiere.objects.get(subject_code=row[3].strip())
        comp_eval, ce_created = CompetencesEvaluees.objects.update_or_create(matiere=matiere, competence=comp)

        print(f"Competence : {comp} has been created : {c_created} and Eval {comp_eval} has also been created : {ce_created}")


def upload_data_from_excel_objectifs_resultats(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(io.BytesIO(file_path))
    # workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    reshaped_data = []
    # Iterate through rows and create new instances
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            student = Student.objects.get(fiche=row[0])
        except Student.DoesNotExist:
            # print(row[1])  # Assuming 'Nom de l'élève' is in the second column
            continue

        matiere = Matiere.objects.get(subject_code=row[3])
        competence = Competence.objects.get(nom=row[7])
        competence_evaluee = CompetencesEvaluees.objects.get(matiere=matiere, competence=competence)
        new_assessement = row[8]  # Assuming 'Description' is in the fourth column
        
        if new_assessement:
            is_most_recent_assessement = True
        else:
            is_most_recent_assessement  = False
            
        # Iterate through grade columns and add non-null grades to the reshaped data
        for semester, grade in enumerate(row[10:18], start=1):
            if grade is not None:
                lookup_attributes = {"etudiant": student, 
                                     "competence_evaluee": competence_evaluee, 
                                     "etape": semester
                                     }
                update_attributes = {"note": grade}
                try:
                    e, created = Evaluation.objects.update_or_create(**lookup_attributes, defaults=update_attributes)
                    print(e, created)
                except:
                    print("********************** ERRROR ")